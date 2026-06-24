"""
Module: export_excel.py
Purpose: Generate formatted Excel report from pipeline results.
         Returns BytesIO buffer — no file I/O, no Streamlit dependency.
"""
from io import BytesIO
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── Color constants ──────────────────────────────────────────────────────────
_HEADER_FILL = PatternFill(start_color="1A365D", end_color="1A365D", fill_type="solid")
_HEADER_FONT = Font(name="Segoe UI", bold=True, color="FFFFFF", size=11)
_BOD10_FILL = PatternFill(start_color="FEF9C3", end_color="FEF9C3", fill_type="solid")
_BOD5_FILL = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
_LABEL_FONT = Font(name="Segoe UI", bold=True, size=11)
_VALUE_FONT = Font(name="Segoe UI", size=11)
_THIN_BORDER = Border(
    left=Side(style="thin", color="E2E8F0"),
    right=Side(style="thin", color="E2E8F0"),
    top=Side(style="thin", color="E2E8F0"),
    bottom=Side(style="thin", color="E2E8F0"),
)


def generate_excel_report(
    sample_name: str,
    peaks_df: pd.DataFrame,
    classification: str,
    probability: float,
    toxicity_pct=None,
    stage1_tag=None,
    stage1_ddo_avg=None,
    stage2_tag=None,
    stage2_ddo_avg=None,
    signal_points: int = 0,
    do_min: float = 0.0,
    do_max: float = 0.0,
    bod_phase1=None,
    bod_phase2=None,
) -> BytesIO:
    """
    Build a formatted 2-sheet Excel report and return it as an in-memory buffer.

    Parameters
    ----------
    sample_name : str
        Filename-derived sample identifier.
    peaks_df : pd.DataFrame
        Columns: No.peak, Tag, Doin (mV), DOmin (mV), DDO (mV), Sample Name.
    classification : str
        Predicted class label (e.g. "GGA").
    probability : float
        Max class probability, 0-1 scale.
    toxicity_pct : float | None
        Toxicity percentage, or None if unavailable.
    stage1_tag, stage2_tag : str | None
        Tag names for stage 1 / stage 2.
    stage1_ddo_avg, stage2_ddo_avg : float | None
        Average DDO for each stage.
    signal_points : int
        Number of raw data points in the signal.
    do_min, do_max : float
        Min/max DO values from raw signal.
    bod_phase1 : float | None
        BOD Phase 1 value (mg/L), or None if unavailable.
    bod_phase2 : float | None
        BOD Phase 2 value (mg/L), or None if unavailable.

    Returns
    -------
    BytesIO
        In-memory .xlsx file ready for st.download_button.
    """
    wb = Workbook()

    # ── Sheet 1: Summary ─────────────────────────────────────────────────
    ws_sum = wb.active
    ws_sum.title = "Summary"
    ws_sum.sheet_properties.tabColor = "4CAF50"

    # Title row
    ws_sum.merge_cells("A1:B1")
    title_cell = ws_sum["A1"]
    title_cell.value = "VHL Biology — Analysis Report"
    title_cell.font = Font(name="Segoe UI", bold=True, size=14, color="1A365D")
    title_cell.alignment = Alignment(horizontal="center")

    # Summary rows: (label, value) pairs starting at row 3
    tag_counts = peaks_df["Tag"].value_counts()
    bod_breakdown = ", ".join(f"{count} {tag}" for tag, count in tag_counts.items())

    summary_rows = [
        ("Sample Name / Ten mau", sample_name),
        ("Signal Points / So diem tin hieu", signal_points),
        ("DO Range (mV)", f"{do_min:.2f} – {do_max:.2f}"),
        ("Peak Count / So peak", f"{len(peaks_df)} ({bod_breakdown})"),
        ("Classification / Phan loai", f"{classification} ({probability * 100:.1f}%)"),
        ("Toxicity / Do doc (%)", f"{toxicity_pct}%" if toxicity_pct is not None else "N/A"),
        ("Stage 1", f"{stage1_tag} — DDO avg {stage1_ddo_avg:.2f} mV" if stage1_tag and stage1_ddo_avg is not None else "N/A"),
        ("Stage 2", f"{stage2_tag} — DDO avg {stage2_ddo_avg:.2f} mV" if stage2_tag and stage2_ddo_avg is not None else "N/A"),
        ("DDO Range (mV)", f"{peaks_df['DDO (mV)'].min():.2f} – {peaks_df['DDO (mV)'].max():.2f}"),
        ("Analysis Date / Ngay phan tich", datetime.now().strftime("%Y-%m-%d %H:%M")),
    ]
    if bod_phase1 is not None:
        summary_rows.insert(5, ("BOD Phase 2 (mg/L)", f"{bod_phase2:.3f}"))
        summary_rows.insert(5, ("BOD Phase 1 (mg/L)", f"{bod_phase1:.3f}"))

    for i, (label, value) in enumerate(summary_rows, start=3):
        cell_label = ws_sum.cell(row=i, column=1, value=label)
        cell_label.font = _LABEL_FONT
        cell_label.border = _THIN_BORDER

        cell_value = ws_sum.cell(row=i, column=2, value=value)
        cell_value.font = _VALUE_FONT
        cell_value.border = _THIN_BORDER

    ws_sum.column_dimensions["A"].width = 36
    ws_sum.column_dimensions["B"].width = 48

    # ── Sheet 2: Peaks ───────────────────────────────────────────────────
    ws_peaks = wb.create_sheet("Peaks")
    ws_peaks.sheet_properties.tabColor = "2196F3"

    columns = ["No.peak", "Tag", "Doin (mV)", "DOmin (mV)", "DDO (mV)"]
    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws_peaks.cell(row=1, column=col_idx, value=col_name)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = _THIN_BORDER

    for row_idx, (_, row) in enumerate(peaks_df.iterrows(), start=2):
        tag_val = str(row.get("Tag", "")).strip()
        row_fill = _BOD10_FILL if "BOD10" in tag_val else _BOD5_FILL

        for col_idx, col_name in enumerate(columns, start=1):
            val = row.get(col_name, "")
            # Format numeric values to 2 decimal places
            if col_name in ("Doin (mV)", "DOmin (mV)", "DDO (mV)"):
                try:
                    val = round(float(val), 2)
                except (ValueError, TypeError):
                    pass
            cell = ws_peaks.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = row_fill
            cell.border = _THIN_BORDER
            if col_name in ("Doin (mV)", "DOmin (mV)", "DDO (mV)"):
                cell.number_format = "0.00"
                cell.alignment = Alignment(horizontal="right")
            else:
                cell.alignment = Alignment(horizontal="center")

    # Auto-width for peaks columns
    for col_idx in range(1, len(columns) + 1):
        ws_peaks.column_dimensions[get_column_letter(col_idx)].width = 16

    # ── Save to buffer ───────────────────────────────────────────────────
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
