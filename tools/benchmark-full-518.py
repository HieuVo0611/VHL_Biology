"""
Benchmark script: runs 518 labeled samples (GGA + GGA-metal) through full pipeline.
Measures per-step timing and collects classification + phase detection results.
Output: plans/reports/benchmark-260622-full-518.xlsx
"""
import sys, os, time
import pandas as pd
from tqdm import tqdm
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FILL_CORRECT = PatternFill("solid", fgColor="C6EFCE")
FILL_WRONG   = PatternFill("solid", fgColor="FFC7CE")
FILL_ERROR   = PatternFill("solid", fgColor="FFEB9C")
FILL_HEADER  = PatternFill("solid", fgColor="1F4E79")
FONT_HEADER  = Font(color="FFFFFF", bold=True, size=10)
FONT_BOLD    = Font(bold=True)
ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
THIN_BORDER  = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from src.utils import extract_peaks_from_txt, catboost_inference_from_csv
from src.phase_detector import update_phase_tags

MODEL_PATH         = "model/catboost_model.cbm"
LABEL_ENCODER_PATH = "model/label_encoder_classes.npy"
GGA_DIR            = "data/GGA/File txt"
METAL_DIR          = "data/GGA-metal/File txt"
OUTPUT_PATH        = "plans/reports/benchmark-260622-full-518.xlsx"


def _header_row(ws, cols):
    for col_idx, col_name in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill      = FILL_HEADER
        cell.font      = FONT_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border    = THIN_BORDER


def _write_sheet2(wb, records):
    ws = wb.create_sheet("Per-Sample Detail")
    cols = [
        "sample_name", "true_label", "pred_label", "correct",
        "confidence_%", "n_peaks",
        "n_phase1", "n_transition", "n_phase2",
        "t_extract_s", "t_classify_s", "t_phase_s", "t_total_s",
        "error_msg",
    ]
    _header_row(ws, cols)

    sorted_records = sorted(records, key=lambda r: r["confidence_pct"])

    for row_idx, rec in enumerate(sorted_records, start=2):
        values = [
            rec["sample_name"], rec["true_label"], rec["pred_label"],
            "✓" if rec["correct"] else ("ERR" if rec["error_msg"] else "✗"),
            rec["confidence_pct"], rec["n_peaks"],
            rec["n_phase1"], rec["n_transition"], rec["n_phase2"],
            rec["t_extract"], rec["t_classify"], rec["t_phase"], rec["t_total"],
            rec["error_msg"],
        ]
        if rec["error_msg"]:
            fill = FILL_ERROR
        elif rec["correct"]:
            fill = FILL_CORRECT
        else:
            fill = FILL_WRONG

        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill      = fill
            cell.border    = THIN_BORDER
            cell.alignment = Alignment(vertical="center")

    widths = [45, 12, 12, 8, 14, 8, 10, 14, 10, 12, 13, 11, 11, 40]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}1"


def _write_sheet3(wb, records):
    """Confidence distribution — shows accuracy per confidence bucket."""
    ws = wb.create_sheet("Distribution")

    cols = ["Confidence Range", "Total Samples", "Correct", "Wrong", "Accuracy %"]
    _header_row(ws, cols)

    buckets = [(0,50), (50,60), (60,70), (70,80), (80,90), (90,101)]
    labels  = ["0–50%", "50–60%", "60–70%", "70–80%", "80–90%", "90–100%"]

    df = pd.DataFrame(records)
    df_valid = df[df["error_msg"] == ""]

    for row_idx, ((lo, hi), label) in enumerate(zip(buckets, labels), start=2):
        subset  = df_valid[(df_valid["confidence_pct"] >= lo) & (df_valid["confidence_pct"] < hi)]
        total   = len(subset)
        correct = int(subset["correct"].sum()) if total > 0 else 0
        wrong   = total - correct
        acc     = round(correct / total * 100, 1) if total > 0 else 0.0

        if acc >= 80:
            fill = PatternFill("solid", fgColor="C6EFCE")
        elif acc >= 60:
            fill = PatternFill("solid", fgColor="FFEB9C")
        else:
            fill = PatternFill("solid", fgColor="FFC7CE")

        for col_idx, val in enumerate([label, total, correct, wrong, acc], start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill      = fill
            cell.border    = THIN_BORDER
            cell.alignment = ALIGN_CENTER

    # Totals row
    total_all   = len(df_valid)
    correct_all = int(df_valid["correct"].sum()) if total_all > 0 else 0
    row = ws.max_row + 1
    for col_idx, val in enumerate(
        ["TOTAL", total_all, correct_all, total_all - correct_all,
         round(correct_all / total_all * 100, 1) if total_all > 0 else 0],
        start=1
    ):
        cell = ws.cell(row=row, column=col_idx, value=val)
        cell.font      = FONT_BOLD
        cell.border    = THIN_BORDER
        cell.alignment = ALIGN_CENTER

    for i, w in enumerate([16, 16, 12, 12, 14], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _write_sheet1(wb, records):
    """Summary sheet — aggregate accuracy and timing stats."""
    ws = wb.create_sheet("Summary", 0)

    df = pd.DataFrame(records)
    df_valid   = df[df["error_msg"] == ""]
    df_gga     = df_valid[df_valid["true_label"] == "gga"]
    df_metal   = df_valid[df_valid["true_label"] == "gga-metal"]
    df_correct = df_valid[df_valid["correct"] == True]
    df_wrong   = df_valid[df_valid["correct"] == False]

    n_total   = len(df_valid)
    n_correct = int(df_valid["correct"].sum())
    n_errors  = len(df[df["error_msg"] != ""])

    def _row(ws, row, label, value, fill=None):
        lc = ws.cell(row=row, column=1, value=label)
        vc = ws.cell(row=row, column=2, value=value)
        lc.font = FONT_BOLD
        lc.border = vc.border = THIN_BORDER
        lc.alignment = vc.alignment = Alignment(vertical="center")
        if fill:
            lc.fill = vc.fill = fill

    # Title row
    ws.merge_cells("A1:B1")
    title = ws.cell(row=1, column=1, value="VHL Biology — Full 518-Sample Benchmark")
    title.font      = Font(color="FFFFFF", bold=True, size=14)
    title.fill      = FILL_HEADER
    title.alignment = ALIGN_CENTER

    row = 2
    _row(ws, row, "Total samples processed", n_total);                       row += 1
    _row(ws, row, "Errors / skipped",         n_errors);                     row += 1
    _row(ws, row, "Correct",  n_correct,       FILL_CORRECT);                row += 1
    _row(ws, row, "Wrong",    n_total - n_correct, FILL_WRONG);              row += 1
    _row(ws, row, "Overall Accuracy %",
         f"{n_correct/n_total*100:.1f}%" if n_total > 0 else "N/A");        row += 1

    row += 1  # spacer
    _row(ws, row, "--- Per-class ---", "");                                   row += 1

    gga_acc   = df_gga["correct"].mean()   * 100 if len(df_gga)   > 0 else 0
    metal_acc = df_metal["correct"].mean() * 100 if len(df_metal) > 0 else 0
    _row(ws, row, f"GGA Accuracy   ({len(df_gga)} samples)",
         f"{gga_acc:.1f}%");                                                  row += 1
    _row(ws, row, f"Metal Accuracy ({len(df_metal)} samples)",
         f"{metal_acc:.1f}%");                                                row += 1

    row += 1  # spacer
    _row(ws, row, "--- Confidence ---", "");                                  row += 1
    avg_conf_correct = df_correct["confidence_pct"].mean() if len(df_correct) > 0 else 0
    avg_conf_wrong   = df_wrong["confidence_pct"].mean()   if len(df_wrong)   > 0 else 0
    _row(ws, row, "Avg confidence (correct)", f"{avg_conf_correct:.1f}%");   row += 1
    _row(ws, row, "Avg confidence (wrong)",   f"{avg_conf_wrong:.1f}%");     row += 1

    row += 1  # spacer
    _row(ws, row, "--- Avg Timing (per sample) ---", "");                    row += 1
    for step, col in [("Extract",  "t_extract"),
                      ("Classify", "t_classify"),
                      ("Phase",    "t_phase"),
                      ("Total",    "t_total")]:
        avg = df_valid[col].mean() if len(df_valid) > 0 else 0
        _row(ws, row, f"Avg {step} time", f"{avg:.3f}s");                   row += 1

    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 20


def find_txt_files(base_dir):
    """Recursively find all .txt files under base_dir."""
    results = []
    for root, dirs, files in os.walk(base_dir):
        for f in files:
            if f.lower().endswith(".txt"):
                results.append(os.path.join(root, f))
    return sorted(results)


def run_benchmark():
    gga_files   = [(f, "gga")       for f in find_txt_files(GGA_DIR)]
    metal_files = [(f, "gga-metal") for f in find_txt_files(METAL_DIR)]
    all_files   = gga_files + metal_files
    print(f"Found {len(gga_files)} GGA + {len(metal_files)} GGA-metal = {len(all_files)} total\n")

    records = []
    for fpath, true_label in tqdm(all_files, desc="Benchmarking"):
        sample_name = os.path.basename(fpath).replace(".txt", "")
        rec = dict(
            sample_name=sample_name, true_label=true_label,
            pred_label="", correct=False, confidence_pct=0.0,
            n_peaks=0, n_phase1=0, n_transition=0, n_phase2=0,
            t_extract=0.0, t_classify=0.0, t_phase=0.0, t_total=0.0,
            error_msg="",
        )
        t_start = time.time()
        try:
            # Step 1: extract peaks
            t0 = time.time()
            peaks_df = extract_peaks_from_txt(fpath)
            rec["t_extract"] = round(time.time() - t0, 3)
            rec["n_peaks"]   = len(peaks_df)

            if len(peaks_df) == 0:
                rec["error_msg"] = "0 peaks extracted"
                rec["t_total"]   = round(time.time() - t_start, 3)
                records.append(rec)
                continue

            # Step 2: classify
            t0 = time.time()
            preds = catboost_inference_from_csv(peaks_df, MODEL_PATH, LABEL_ENCODER_PATH)
            _, pred_label, pred_proba = preds[0]
            pred_label = str(pred_label).strip().lower()
            rec["t_classify"]     = round(time.time() - t0, 3)
            rec["pred_label"]     = pred_label
            rec["correct"]        = (pred_label == true_label)
            rec["confidence_pct"] = round(max(pred_proba) * 100, 1)

            # Step 3: phase detection
            t0 = time.time()
            peaks_df = update_phase_tags(peaks_df, pred_label)
            rec["t_phase"] = round(time.time() - t0, 3)

            tag_counts        = peaks_df["Tag"].str.strip().value_counts()
            rec["n_phase1"]     = int(tag_counts.get("phase1",     0))
            rec["n_transition"] = int(tag_counts.get("transition", 0))
            rec["n_phase2"]     = int(tag_counts.get("phase2",     0))

        except Exception as e:
            rec["error_msg"] = str(e)[:120]

        rec["t_total"] = round(time.time() - t_start, 3)
        records.append(rec)

    return records


if __name__ == "__main__":
    records = run_benchmark()
    print(f"\nCollected {len(records)} records")
    wb = Workbook()
    wb.remove(wb.active)   # remove default empty sheet
    _write_sheet2(wb, records)
    _write_sheet3(wb, records)
    _write_sheet1(wb, records)
    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
    wb.save(OUTPUT_PATH)
    print(f"Saved: {OUTPUT_PATH}")
