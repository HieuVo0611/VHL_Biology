"""
Extract phase-boundary GT from Excel files (yellow=stable, white=transition).
Outputs 3 CSVs to data/phase-gt-{gga,metal,hh}.csv with columns:
  [Sample Name, Sheet Name, Block Index, peak_idx, Tag, Doin, DOmin, DDO, phase_label]
phase_label: 0=phase1 (first yellow block), 1=transition (white middle), 2=phase2 (last yellow block)

Sheets may contain MULTIPLE Doin blocks side-by-side (each is a separate sample).
Headers may be in row 2 or row 3; data starts on the row immediately after.
Blocks with no color mix (all-yellow or all-white) are skipped - cannot derive GT.
"""
import sys
import os
import openpyxl
import pandas as pd

sys.stdout.reconfigure(encoding='utf-8')

YELLOW_VARIANTS = {'FFFFFF00', 'FFFEF2CB'}

EXCEL_SOURCES = {
    'gga':   'data/BOD-Hieu/GGA/File excel/Excel-GGA.xlsx',
    'metal': 'data/BOD-Hieu/GGA-metal/File Excel/Excel-GGA-metal.xlsx',
    'hh':    'data/BOD-Hieu/GGA-metal/File Excel/Excel-GGA-metal HH.xlsx',
}

def detect_doin_columns(ws):
    """
    Scan rows 2 and 3 for headers containing 'Doin'. Returns list of (header_row, col) tuples
    for every Doin column found. Data row starts at header_row + 1.
    """
    found = []
    for hr in (2, 3):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(hr, c).value
            if v and 'Doin' in str(v):
                found.append((hr, c))
    return found

def detect_sample_name_for_block(ws, doin_col):
    """
    Sample name is typically the full string containing 'Q=' located in row 1 (or row 2)
    near the same column as the Doin block. Strategy: scan row 1 and row 2 across all columns,
    pick the 'Q='-containing string whose column is closest to doin_col (preferring same column
    or to the left of doin_col since blocks are usually labeled at their starting column).
    Fallback to sheet title.
    """
    candidates = []  # (distance, col, value)
    for r in (1, 2):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and 'Q=' in v:
                candidates.append((abs(c - doin_col), c, v.strip()))
    if not candidates:
        return ws.title
    candidates.sort(key=lambda x: x[0])
    return candidates[0][2]

def labels_from_colors(colors):
    """
    colors: list of 'Y'/'W' per peak in order.
    Returns: list of phase_label (0/1/2), or None if pattern is invalid (no Y-W-Y).
    Valid pattern: Y+ W+ Y+ (first yellow block, middle white block, last yellow block).
    """
    if not colors or 'Y' not in colors or 'W' not in colors:
        return None
    first_w = colors.index('W')
    last_w = len(colors) - 1 - colors[::-1].index('W')
    if any(c != 'Y' for c in colors[:first_w]):
        return None
    if any(c != 'Y' for c in colors[last_w + 1:]):
        return None
    labels = []
    for i, c in enumerate(colors):
        if i < first_w:
            labels.append(0)
        elif i > last_w:
            labels.append(2)
        else:
            labels.append(1)
    return labels

def extract_block(ws, sheet_name, header_row, doin_col, block_idx):
    """Extract rows from one Doin block. Returns list of dicts or [] if not extractable."""
    sample_name = detect_sample_name_for_block(ws, doin_col)
    data_start = header_row + 1
    rows_raw = []
    for r in range(data_start, ws.max_row + 1):
        cell = ws.cell(r, doin_col)
        if cell.value is None or not isinstance(cell.value, (int, float)):
            continue
        fill = cell.fill.fgColor.rgb if cell.fill and cell.fill.fgColor else '00000000'
        color = 'Y' if fill in YELLOW_VARIANTS else 'W'
        tag = ws.cell(r, doin_col - 1).value
        no_peak = ws.cell(r, doin_col + 1).value
        domin = ws.cell(r, doin_col + 2).value
        ddo = ws.cell(r, doin_col + 3).value
        rows_raw.append({
            'Sample Name': sample_name,
            'Sheet Name': sheet_name,
            'Block Index': block_idx,
            'peak_idx': len(rows_raw),
            'Tag': tag,
            'Doin': cell.value,
            'No.peak': no_peak,
            'DOmin': domin,
            'DDO': ddo,
            'color': color,
        })
    if not rows_raw:
        return []
    colors = [r['color'] for r in rows_raw]
    labels = labels_from_colors(colors)
    if labels is None:
        return []
    for r, lab in zip(rows_raw, labels):
        r['phase_label'] = lab
        del r['color']
    return rows_raw

def extract_sheet(ws, sheet_name):
    """Extract all Doin blocks from a sheet. Returns (rows, n_blocks_used, n_blocks_skipped)."""
    blocks = detect_doin_columns(ws)
    if not blocks:
        return [], 0, 0
    all_rows = []
    used = 0
    skipped = 0
    for idx, (hr, col) in enumerate(blocks):
        rows = extract_block(ws, sheet_name, hr, col, idx)
        if rows:
            all_rows.extend(rows)
            used += 1
        else:
            skipped += 1
    return all_rows, used, skipped

def main():
    os.makedirs('data', exist_ok=True)
    for kind, path in EXCEL_SOURCES.items():
        if not os.path.exists(path):
            print(f'[SKIP] {path} not found')
            continue
        print(f'[LOAD] {path}')
        wb = openpyxl.load_workbook(path, data_only=True)
        all_rows = []
        sheets_with_data = 0
        total_blocks_used = 0
        total_blocks_skipped = 0
        for sn in wb.sheetnames:
            rows, used, skipped = extract_sheet(wb[sn], sn)
            if rows:
                all_rows.extend(rows)
                sheets_with_data += 1
            total_blocks_used += used
            total_blocks_skipped += skipped
        out_path = f'data/phase-gt-{kind}.csv'
        df = pd.DataFrame(all_rows)
        df.to_csv(out_path, index=False, encoding='utf-8')
        print(f'  -> {out_path}: {len(df)} peaks from {total_blocks_used} blocks '
              f'across {sheets_with_data} sheets ({total_blocks_skipped} blocks skipped)')

if __name__ == '__main__':
    main()
