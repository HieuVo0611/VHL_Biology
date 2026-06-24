import pytest
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook

from src.utils import calculate_bod_from_calibration
from src.export_excel import generate_excel_report


def test_basic_linear_fit():
    # 2 calibration points: BOD=20 → DDO=11.3, BOD=15 → DDO=9.13
    # slope a = (9.13 - 11.3) / (15 - 20) = -2.17 / -5 = 0.434
    # intercept b = 11.3 - 0.434 * 20 = 11.3 - 8.68 = 2.62
    # BOD from DDO=10: (10 - 2.62) / 0.434 = 17.004...
    # BOD from DDO=6:  (6  - 2.62) / 0.434 = 7.788...
    result = calculate_bod_from_calibration(
        ddo_phase1=10.0, ddo_phase2=6.0,
        bod1_cal=20.0, ddo1_cal=11.3,
        bod2_cal=15.0, ddo2_cal=9.13,
    )
    assert abs(result["a"] - 0.434) < 0.001
    assert abs(result["b"] - 2.62)  < 0.01
    assert abs(result["bod_phase1"] - 17.004) < 0.01
    assert abs(result["bod_phase2"] - 7.788)  < 0.01


def test_raises_on_equal_bod_calibration():
    with pytest.raises(ValueError, match="khác nhau"):
        calculate_bod_from_calibration(
            ddo_phase1=10.0, ddo_phase2=6.0,
            bod1_cal=20.0, ddo1_cal=11.3,
            bod2_cal=20.0, ddo2_cal=9.13,  # same BOD → zero slope
        )


def test_return_keys_present():
    result = calculate_bod_from_calibration(
        ddo_phase1=10.0, ddo_phase2=6.0,
        bod1_cal=20.0, ddo1_cal=11.3,
        bod2_cal=15.0, ddo2_cal=9.13,
    )
    assert set(result.keys()) == {"a", "b", "bod_phase1", "bod_phase2"}


def test_values_rounded_to_3dp():
    result = calculate_bod_from_calibration(
        ddo_phase1=10.0, ddo_phase2=6.0,
        bod1_cal=20.0, ddo1_cal=11.3,
        bod2_cal=15.0, ddo2_cal=9.13,
    )
    # bod_phase1 and bod_phase2 rounded to 3 decimal places
    assert result["bod_phase1"] == round(result["bod_phase1"], 3)
    assert result["bod_phase2"] == round(result["bod_phase2"], 3)


def _make_peaks_df():
    return pd.DataFrame({
        "No.peak": [1, 2, 3, 4],
        "Tag": ["phase1", "phase1", "phase2", "phase2"],
        "Doin (mV)": [260.0, 261.0, 258.0, 259.0],
        "DOmin (mV)": [255.0, 256.0, 254.0, 255.0],
        "DDO (mV)": [5.0, 5.0, 4.0, 4.0],
        "Sample Name": ["test"] * 4,
    })


def test_excel_report_includes_bod_rows():
    buf = generate_excel_report(
        sample_name="test", peaks_df=_make_peaks_df(),
        classification="gga", probability=0.86,
        toxicity_pct=None,
        stage1_tag="phase1", stage1_ddo_avg=5.0,
        stage2_tag="phase2", stage2_ddo_avg=4.0,
        signal_points=1000, do_min=255.0, do_max=265.0,
        bod_phase1=17.004, bod_phase2=7.788,
    )
    wb = load_workbook(BytesIO(buf.read()))
    ws = wb["Summary"]
    labels = [ws.cell(row=r, column=1).value for r in range(3, 20) if ws.cell(row=r, column=1).value]
    assert any("BOD Phase 1" in str(l) for l in labels)
    assert any("BOD Phase 2" in str(l) for l in labels)


def test_excel_report_no_bod_rows_when_none():
    buf = generate_excel_report(
        sample_name="test", peaks_df=_make_peaks_df(),
        classification="gga-metal", probability=0.82,
        toxicity_pct=12.5,
        stage1_tag="phase1", stage1_ddo_avg=5.0,
        stage2_tag="phase2", stage2_ddo_avg=4.0,
        signal_points=1000, do_min=255.0, do_max=265.0,
    )
    wb = load_workbook(BytesIO(buf.read()))
    ws = wb["Summary"]
    labels = [ws.cell(row=r, column=1).value for r in range(3, 20) if ws.cell(row=r, column=1).value]
    assert not any("BOD Phase" in str(l) for l in labels)
